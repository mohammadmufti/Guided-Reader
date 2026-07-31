"""
The gold standard: a hand-checked sample that turns agreement into accuracy.

Every accuracy figure in this project is agreement between two automated
sources (GAPS.md §2). This tool draws a stratified sample of tokens, emits a
review workbook for a human to check against scholarly references, and scores
the verdicts into per-stratum accuracy with confidence intervals.

    python pipeline/gold.py sample --corpus tajrid     # draw + write the sheet
    python pipeline/gold.py score  --corpus tajrid     # read verdicts, report

Design decisions, so they are not re-litigated later:

- **Strata are a strict partition**, assigned in priority order, so every
  token belongs to exactly one and population weights are exact. The suspect
  classes are oversampled on purpose: a uniform draw of 300 would contain
  about nine heuristic tokens and measure nothing about them.
- **The sample is deterministic** — fixed seed over candidates sorted by
  (record id, token index). The same payload always yields the same sample.
- **The sheet snapshots what shipped.** A verdict certifies the value the
  reader saw at sample time. test_gold.py fails loudly if the payload drifts
  under a drawn sample, because at that point the verdicts no longer describe
  the shipping build and the report must be re-scored.
- **Witnessed vowelling gets a small stratum, not a large one.** For vowels
  the aligned Bukhārī is already ground truth at scale (DIACRITISATION.md §1);
  the hand-check there audits the *alignment*, not the vowels. The large
  strata go where no ground truth exists at all: roots, lemmas, POS
  everywhere, and vowelling on the unwitnessed tiers.
"""

from __future__ import annotations

import argparse
import glob
import json
import math
import random
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
DATA = ROOT / "web" / "public" / "data"
GOLD = ROOT / "pipeline" / "gold"

SEED = 20260730  # the date Phase 1 opened; never change once verdicts exist

# stratum -> (n to draw). Assignment below is in THIS order; first match wins.
STRATA = {
    "override": 30,        # contextRoot fired — audits the geminate→hollow rule
    "suspect": 60,         # analysers disagree or the morphology is flagged
    "heuristic_case": 50,  # vowel from the case-agreement rule (no ground truth)
    "heuristic_mfv": 50,   # vowel from most-frequent-reading (no ground truth)
    "lexicon_unique": 60,  # single lexicon candidate, unwitnessed
    "witnessed": 50,       # aligned to the Bukhārī witness — audits alignment
}

VERDICT_COLS = ["vowels_ok", "root_ok", "lemma_ok", "pos_ok"]


# ---------------------------------------------------------------- payload --

def load_payload():
    surface: dict = {}
    for f in sorted(glob.glob(str(DATA / "lex" / "surface-*.json"))):
        surface.update(json.loads(Path(f).read_text(encoding="utf-8")))
    records = []
    for f in sorted(glob.glob(str(DATA / "hadith" / "matn-*.json"))):
        records.append(json.loads(Path(f).read_text(encoding="utf-8")))
    if not surface or not records:
        sys.exit(f"payload not found under {DATA} — run the pipeline first")
    build_id = json.loads((DATA / "index.json").read_text(encoding="utf-8")).get("buildId")
    return surface, records, build_id


def stratum_of(tok: dict, entry: dict | None) -> str | None:
    """First match wins; must mirror the STRATA ordering."""
    if not tok.get("clickable") or not tok.get("matchId") or entry is None:
        return None  # unbound: nothing shown, nothing to verify
    if tok.get("contextRoot"):
        return "override"
    if entry.get("rootDisputed") or entry.get("morphSuspect") \
            or entry.get("pos_agreement") == "disagree":
        return "suspect"
    if tok.get("binding") == "heuristic":
        return "heuristic_case" if tok.get("confidence") == "medium" else "heuristic_mfv"
    if tok.get("binding") == "unique":
        return "lexicon_unique"
    if tok.get("binding") == "aligned":
        return "witnessed"
    return None


def flat_gloss(g) -> str:
    """gloss is structured ({senses, before, after, ...}); render the senses."""
    if not g:
        return ""
    if isinstance(g, str):
        return g
    if isinstance(g, dict):
        return "; ".join(s for s in (g.get("senses") or []) if s)[:80]
    return str(g)[:80]


def snapshot(tok: dict, entry: dict, rec: dict) -> dict:
    toks = rec["tokens"]
    i = tok["i"]
    before = " ".join(t["raw"] for t in toks[max(0, i - 4): i])
    after = " ".join(t["raw"] for t in toks[i + 1: i + 5])
    return {
        "record": rec["id"],
        "number": rec["number"],
        "i": i,
        "raw": tok["raw"],
        "shown_vowels": tok["surface"],
        # Mirrors the panel's precedence exactly, rootPreferAnalysed included
        # — a verdict on "shown_root" must judge what the reader actually saw.
        "shown_root": tok.get("contextRoot")
        or (entry.get("rootPreferAnalysed")
            and (entry.get("analysed") or {}).get("root"))
        or entry.get("root")
        or (entry.get("analysed") or {}).get("root")
        or "",
        "shown_lemma": entry.get("lemma") or "",
        "shown_pos": entry.get("pos") or "",
        "gloss": flat_gloss(entry.get("gloss")),
        "binding": tok.get("binding"),
        "confidence": tok.get("confidence"),
        "before": before,
        "after": after,
    }


# ----------------------------------------------------------------- sample --

def cmd_sample(corpus: str) -> int:
    surface, records, build_id = load_payload()

    pools: dict[str, list[dict]] = {k: [] for k in STRATA}
    for rec in records:
        if not rec.get("number"):
            continue
        for tok in rec["tokens"]:
            entry = surface.get(tok.get("matchId") or "")
            s = stratum_of(tok, entry)
            if s:
                pools[s].append(snapshot(tok, entry, rec))

    rng = random.Random(SEED)
    rows, populations = [], {}
    for name, n in STRATA.items():
        pool = sorted(pools[name], key=lambda r: (r["record"], r["i"]))
        populations[name] = len(pool)
        if len(pool) < n:
            print(f"  note: stratum {name} has only {len(pool)} tokens; taking all")
        take = pool if len(pool) <= n else rng.sample(pool, n)
        for r in sorted(take, key=lambda r: (r["record"], r["i"])):
            rows.append({"stratum": name, **r})
    for k, r in enumerate(rows, 1):
        r["id"] = f"G{k:03d}"

    out = GOLD / corpus
    out.mkdir(parents=True, exist_ok=True)
    doc = {"schemaVersion": 1, "corpus": corpus, "seed": SEED,
           "buildId": build_id, "populations": populations, "rows": rows}
    (out / "sample.json").write_text(
        json.dumps(doc, ensure_ascii=False, indent=1), encoding="utf-8")
    write_sheet(out / "review.xlsx", doc)
    print(f"{len(rows)} tokens across {len(STRATA)} strata")
    for k in STRATA:
        print(f"  {k}: {sum(1 for r in rows if r['stratum'] == k)}"
              f"  (population {populations[k]:,})")
    print(f"-> {out / 'sample.json'}\n-> {out / 'review.xlsx'}")
    return 0


# ------------------------------------------------------------------ sheet --

def write_sheet(path: Path, doc: dict) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    URL = "https://mohammadmufti.com/Guided-Reader/hadith/{n}?w={i}"
    HEAD = ["id", "stratum", "hadith", "word", "open in reader",
            "context after", "WORD AS SHOWN", "context before", "unvocalised",
            "shown root", "shown lemma", "shown POS", "gloss",
            "vowels_ok", "root_ok", "lemma_ok", "pos_ok",
            "corrected vowels", "corrected root", "notes"]
    ARABIC = {6, 7, 8, 9, 10, 11, 18, 19}   # 1-based columns holding Arabic
    FILL_IN = {14, 15, 16, 17, 18, 19, 20}  # 1-based columns the checker edits

    wb = Workbook()

    # --- the legend sheet -------------------------------------------------
    ws = wb.active
    ws.title = "READ ME"
    ws.column_dimensions["A"].width = 100
    lines = [
        "Gold-standard review sheet — how to fill it in",
        "",
        "Check each word IN CONTEXT against a reliable reference (a printed",
        "Bukhārī for witnessed rows; dictionaries and your own grammar for the",
        "rest). The 'open in reader' link shows the word live, in place.",
        "",
        "Fill ONLY the yellow columns, on the Sample sheet:",
        "  vowels_ok / root_ok / lemma_ok / pos_ok — pick yes, no, or unsure.",
        "    'yes' means the shown value is correct FOR THIS OCCURRENCE.",
        "    'unsure' is an honest answer; it is excluded from accuracy, not",
        "    counted against it. Prefer it to a guess.",
        "  corrected vowels / corrected root — only when the verdict is 'no',",
        "    write what it should have been.",
        "  notes — anything the numbers cannot carry.",
        "",
        "Row G000 is a worked example. Do not edit it; it is excluded from",
        "scoring. Everything else about the sheet (shown values, strata) is a",
        "snapshot of the shipping build and must not be edited.",
        "",
        f"Sample of {len(doc['rows'])} tokens, seed {doc['seed']}, "
        f"build {doc['buildId']}.",
        "When done: save, commit this file, and run pipeline/gold.py score.",
    ]
    for r, text in enumerate(lines, 1):
        c = ws.cell(row=r, column=1, value=text)
        c.font = Font(name="Arial", bold=(r == 1), size=14 if r == 1 else 11)

    # --- the sample sheet -------------------------------------------------
    ws = wb.create_sheet("Sample")
    ws.sheet_view.rightToLeft = True  # Arabic-first reading order
    yellow = PatternFill("solid", fgColor="FFF2A0")
    grey = Font(name="Arial", color="666666", size=10)
    arabic_font = Font(name="Arial", size=14)

    for col, h in enumerate(HEAD, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = Font(name="Arial", bold=True)
        if col in FILL_IN:
            c.fill = yellow
    ws.freeze_panes = "A2"

    dv = DataValidation(type="list", formula1='"yes,no,unsure"', allow_blank=True)
    ws.add_data_validation(dv)

    # worked example row, excluded from scoring by its id
    example = ["G000", "example", 1, 3, "(example)", "…قال حدثنا", "الأَعْمَالُ",
               "إنما…", "الاعمال", "عمل", "عَمَل", "noun", "deeds",
               "yes", "yes", "yes", "yes", "", "", "checked against print ed."]
    for col, v in enumerate(example, 1):
        c = ws.cell(row=2, column=col, value=v)
        c.font = grey

    for r, row in enumerate(doc["rows"], 3):
        link = URL.format(n=row["number"], i=row["i"])
        vals = [row["id"], row["stratum"], row["number"], row["i"], "open",
                row["after"], row["shown_vowels"], row["before"], row["raw"],
                row["shown_root"], row["shown_lemma"], row["shown_pos"],
                row["gloss"], "", "", "", "", "", "", ""]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=col, value=v)
            c.font = arabic_font if col in ARABIC else Font(name="Arial")
            if col == 5:
                c.hyperlink = link
                c.font = Font(name="Arial", color="0563C1", underline="single")
            if col in FILL_IN:
                c.fill = yellow
        for col in (14, 15, 16, 17):
            dv.add(ws.cell(row=r, column=col))

    widths = [6, 14, 8, 6, 12, 26, 18, 26, 12, 10, 12, 10, 22,
              10, 9, 9, 8, 16, 12, 24]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.alignment = Alignment(vertical="center")

    wb.save(path)


# ------------------------------------------------------------------ score --

def wilson(k: int, n: int) -> tuple[float, float, float]:
    """point estimate and 95% Wilson interval, as percentages"""
    if n == 0:
        return float("nan"), float("nan"), float("nan")
    z = 1.96
    p = k / n
    d = 1 + z * z / n
    centre = (p + z * z / (2 * n)) / d
    half = z * math.sqrt(p * (1 - p) / n + z * z / (4 * n * n)) / d
    return 100 * p, 100 * (centre - half), 100 * (centre + half)


def cmd_score(corpus: str) -> int:
    from openpyxl import load_workbook

    out = GOLD / corpus
    doc = json.loads((out / "sample.json").read_text(encoding="utf-8"))
    wb = load_workbook(out / "review.xlsx", data_only=True)
    ws = wb["Sample"]

    verdicts: dict[str, dict] = {}
    head = [c.value for c in ws[1]]
    for row in ws.iter_rows(min_row=2, values_only=True):
        rid = row[0]
        if not rid or rid == "G000":
            continue
        verdicts[rid] = dict(zip(head, row))

    rows = doc["rows"]
    filled = sum(1 for r in rows
                 if any(verdicts.get(r["id"], {}).get(c) for c in VERDICT_COLS))
    if filled == 0:
        print("no verdicts yet — fill the yellow columns in review.xlsx first")
        return 1

    lines = [f"# Gold standard — {corpus}", "",
             f"Sample of {len(rows)} tokens (seed {doc['seed']}, "
             f"build `{doc['buildId']}`), {filled} with at least one verdict.",
             "",
             "Accuracy is `yes / (yes + no)`; `unsure` and blank are excluded",
             "and reported. 95% Wilson intervals.", ""]

    for field in VERDICT_COLS:
        lines += [f"## {field}", "",
                  "| stratum | n | yes | no | unsure/blank | accuracy | 95% CI |",
                  "|---|--:|--:|--:|--:|--:|--:|"]
        tot_y = tot_n = 0
        for name in STRATA:
            sub = [verdicts.get(r["id"], {}).get(field)
                   for r in rows if r["stratum"] == name]
            y = sum(1 for v in sub if v == "yes")
            n = sum(1 for v in sub if v == "no")
            u = len(sub) - y - n
            tot_y, tot_n = tot_y + y, tot_n + n
            p, lo, hi = wilson(y, y + n)
            acc = f"{p:.1f}%" if y + n else "—"
            ci = f"{lo:.0f}–{hi:.0f}%" if y + n else "—"
            lines.append(f"| {name} | {len(sub)} | {y} | {n} | {u} | {acc} | {ci} |")
        # corpus-weighted estimate from stratum populations
        pops = doc["populations"]
        total_pop = sum(pops.values())
        est = 0.0
        ok = True
        for name in STRATA:
            sub = [verdicts.get(r["id"], {}).get(field)
                   for r in rows if r["stratum"] == name]
            y = sum(1 for v in sub if v == "yes")
            n = sum(1 for v in sub if v == "no")
            if y + n == 0:
                ok = False
                break
            est += (y / (y + n)) * pops[name] / total_pop
        p, lo, hi = wilson(tot_y, tot_y + tot_n)
        lines += ["",
                  f"Corpus-weighted estimate: **{100 * est:.1f}%**" if ok else
                  "Corpus-weighted estimate: — (a stratum has no verdicts)",
                  f"Unweighted across sample: {p:.1f}% ({lo:.0f}–{hi:.0f}%)", ""]

    report = ROOT / "pipeline" / "reports" / "gold.md"
    report.write_text("\n".join(lines), encoding="utf-8")
    print(f"scored {filled}/{len(rows)} -> {report}")
    return 0


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("cmd", choices=["sample", "score"])
    ap.add_argument("--corpus", default="tajrid")
    a = ap.parse_args()
    return cmd_sample(a.corpus) if a.cmd == "sample" else cmd_score(a.corpus)


if __name__ == "__main__":
    sys.exit(main())
